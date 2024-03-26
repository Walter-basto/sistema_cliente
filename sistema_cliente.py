import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import  openpyxl,xlrd  # excell  baixar
import  pathlib       # ficheiro do excell
from openpyxl import Workbook


#aparencia padrao do sistema
ctk.set_appearance_mode("system")
ctk.set_default_color_theme("blue")

class function():
        
    
              
    def submit(self):
        self.ficheiro=pathlib.Path(r"Clientes.xlsx")
        if self.ficheiro.exists():
            pass
        else:
            self.ficheiro=Workbook()
            #self.folha=self.ficheiro.create_sheet("c")
            self.folha=self.ficheiro.active
            self.folha.title="Sistema de Central de Cliente"
            self.folha['A1']="Nome Completo"
            self.folha['B1']="Contato"
            self.folha['C1']="Idade"
            self.folha['D1']="Gênero"
            self.folha['E1']="Endereço"
            self.folha['F1']="Observação"
            self.ficheiro.save(r"Clientes.xlsx") 


        self.name=self.name_entry.get()
        self.contact=self.contact_entry.get()
        self.age=self.age_entry.get()
        self.andress=self.andress_entry.get() 
        self.gender=self.gender_combobox.get()
        self.observation=self.observation_entry.get("0.0","end")
        try:
            if (self.name=="" or self.contact=="" or self.age =="" or self.andress==""):
                messagebox.showerror(title="Sistema",message="ERRO! Preenchar todos os campos!")
            elif len(self.age)>2:
                 messagebox.showwarning(title="Sistema",message="limite de números são dois!")
            elif len(self.contact)>8:
                 messagebox.showwarning(title="Sistema",message="limite de números são oito!")     
            elif not self.contact.isdigit():
                 messagebox.showwarning(title="Sistema",message="somente números !")
            elif not self.age.isdigit():
                 messagebox.showwarning(title="Sistema",message="somente números !")
            else:
          
                 self.ficheiro=openpyxl.load_workbook(r"Clientes.xlsx")
                 #self.folha=self.ficheiro.get_sheet_by_name(r"Sistema de Central de Cliente")
                 self.folha=self.ficheiro.active
                 self.folha.cell(column=1,row=self.folha.max_row+1,value=self.name)
                 self.folha.cell(column=2,row=self.folha.max_row,value=self.contact)
                 self.folha.cell(column=3,row=self.folha.max_row,value=self.age)
                 self.folha.cell(column=4,row=self.folha.max_row,value=self.gender)
                 self.folha.cell(column=5,row=self.folha.max_row,value=self.andress)
                 self.folha.cell(column=6,row=self.folha.max_row,value=self.observation)
                 self.ficheiro.save(r"Clientes.xlsx")  
                 messagebox.showinfo(title="Sistema",message="Dados salvos com sucesso")
        except:
              messagebox.showerror(title="Sistema",message="ERRO! Preenchar todos os campos!")
        
    def clear(self):
        self.name_entry.delete(0,END)
        self.andress_entry.delete(0,END)
        self.age_entry.delete(0,END)
        self.contact_entry.delete(0,END)
        self.observation_entry.delete(0,END)
      
    
class App(ctk.CTk,function):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appearence()
        self.todo_sistema()
     
      
    def layout_config(self):
        self.title("sistema de central de cliente")
        self.geometry("700x500")
        self.resizable(False,False)
        
    
        
    def appearence(self):
        self.lb_apm=ctk.CTkLabel(self,text="tema",fg_color="transparent",text_color=["#000","#fff"])
        self.lb_apm.place(x=50,y=400)
        self.opt_apm=ctk.CTkOptionMenu(self,values=["Light","Dark"],command=self.change_apm)
        self.opt_apm.place(x=50,y=460)
        
    def change_apm(self,nova_aparencia):   
        ctk.set_appearance_mode(nova_aparencia)   
      
        
      
    def todo_sistema(self):
        self.frame=ctk.CTkFrame(self,width=700,height=50,corner_radius=0,bg_color="teal",fg_color="teal")
        self.frame.place(x=0,y=10)
        self.title=ctk.CTkLabel(self.frame,text="Sistema de Gestão do Cliente",font=("Century Gothic bold",24),text_color=["#000","#fff"]) 
        self.title.place(x=210,y=10)
        self.span=ctk.CTkLabel(self,text="Por favor,preencha todos os campos do formulários!",font=("Century Gothic bold",16),text_color=["#000","#fff"])
        self.span.place(x=50,y=70)  
     
       
        self.lb_name=ctk.CTkLabel(self,text="nome completo:",font=("Century Gothic bold",16),text_color=["#000","#fff"]) 
        self.lb_name.place(x=50,y=120)
        self.name_entry=ctk.CTkEntry(self,width=350,placeholder_text="nome completo sem espaço", font=("Century Gothic bold",16),fg_color="transparent")
        self.name_entry.place(x=50,y=150)
        
        self.lb_andress=ctk.CTkLabel(self,text="endereço:",font=("Century Gothic bold",16),text_color=["#000","#fff"])
        self.lb_andress.place(x=50,y=190)
        self.andress_entry=ctk.CTkEntry(self,width=200,font=("Century Gothic bold",16),fg_color="transparent")
        self.andress_entry.place(x=50,y=220)
        
        self.lb_observation=ctk.CTkLabel(self,text="observações:",font=("Century Gothic bold",16),text_color=["#000","#fff"])
        self.lb_observation.place(x=50,y=260)
        self.observation_entry=ctk.CTkTextbox(self,width=470,height=150,font=("arial",18),border_color="#aaa",border_width=2,fg_color="transparent")
        self.observation_entry.place(x=190,y=260) 
        
        
        self.lb_age=ctk.CTkLabel(self,text="idade:",font=("Century Gothic bold",16),text_color=["#000","#fff"])  
        self.lb_age.place(x=300,y=190)
        self.age_entry=ctk.CTkEntry(self,width=150,placeholder_text="dois números",font=("Century Gothic bold",16),fg_color="transparent")
        self.age_entry.place(x=300,y=220)
          
        
        self.lb_contact=ctk.CTkLabel(self,text="contato:",font=("Century Gothic bold",16),text_color=["#000","#fff"]) 
        self.lb_contact.place(x=450,y=120)
        self.contact_entry=ctk.CTkEntry(self,width=200,placeholder_text="oito números",font=("Century Gothic bold",16),fg_color="transparent")
        self.contact_entry.place(x=450,y=150)
        
        self.lb_gender =ctk.CTkLabel(self,text="gênero:",font=("Century Gothic bold",16),text_color=["#000","#fff"])
        self.lb_gender.place(x=500,y=190)
        self.gender_combobox=ctk.CTkComboBox(self,width=150,values=["Masculino","Feminino"],font=("Century Gothic bold",14))
        self.gender_combobox.set("Masculino")
        self.gender_combobox.place(x=500,y=220)
        
        
        self.btn_submit=ctk.CTkButton(self,text="salvar dados".upper(),command=self.submit,fg_color="blue",hover_color="green",corner_radius=20)  
        self.btn_submit.place(x=250,y=460)
        
        self.btn_clear=ctk.CTkButton(self,text="limpar dados".upper(),command=self.clear,fg_color="red",hover_color="green",corner_radius=20)  
        self.btn_clear.place(x=400,y=460)
  
        
if __name__=="__main__":
     app=App()
     app.mainloop()



